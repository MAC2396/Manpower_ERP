import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Helpers ────────────────────────────────────────────────────────────
HEADER_FONT    = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL    = PatternFill('solid', fgColor='1A237E')
TOTAL_FILL     = PatternFill('solid', fgColor='E8EAF6')
TOTAL_FONT     = Font(bold=True, size=11)
CENTER         = Alignment(horizontal='center', vertical='center')
LEFT           = Alignment(horizontal='left',   vertical='center')
RIGHT          = Alignment(horizontal='right',  vertical='center')

def thin_border():
    s = Side(style='thin', color='BBBBBB')
    return Border(left=s, right=s, top=s, bottom=s)

def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = thin_border()

def style_data_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.border    = thin_border()
        cell.alignment = RIGHT if col > 4 else LEFT

def style_total_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = TOTAL_FONT
        cell.fill      = TOTAL_FILL
        cell.border    = thin_border()
        cell.alignment = RIGHT if col > 4 else LEFT

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or '')))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

def to_stream(wb):
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── Salary Register Excel ──────────────────────────────────────────────
def export_salary_to_excel(salaries, month, year):
    wb = Workbook()
    ws = wb.active
    ws.title = f'Salary {month:02d}-{year}'

    # Title
    ws.merge_cells('A1:N1')
    ws['A1'] = f'SALARY REGISTER — {month:02d}/{year}'
    ws['A1'].font      = Font(bold=True, size=14, color='1A237E')
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 30

    # Headers
    headers = [
        '#', 'Worker Name', 'Company', 'Post', 'Days',
        'Basic', 'DA', 'HRA', 'Special', 'Gross',
        'PF (Emp)', 'ESIC (Emp)', 'Total Ded.', 'Net Pay'
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, len(headers))
    ws.row_dimensions[2].height = 20

    # Data rows
    totals = {k: 0 for k in ['gross','pf','esic','ded','net']}
    for i, s in enumerate(salaries, 1):
        row = i + 2
        ws.cell(row=row, column=1,  value=i)
        ws.cell(row=row, column=2,  value=s.worker.full_name)
        ws.cell(row=row, column=3,  value=s.company.name)
        ws.cell(row=row, column=4,  value=s.worker.post)
        ws.cell(row=row, column=5,  value=s.days_present)
        ws.cell(row=row, column=6,  value=round(s.basic, 2))
        ws.cell(row=row, column=7,  value=round(s.da, 2))
        ws.cell(row=row, column=8,  value=round(s.hra, 2))
        ws.cell(row=row, column=9,  value=round(s.special_allowance, 2))
        ws.cell(row=row, column=10, value=round(s.gross, 2))
        ws.cell(row=row, column=11, value=round(s.pf_employee, 2))
        ws.cell(row=row, column=12, value=round(s.esic_employee, 2))
        ws.cell(row=row, column=13, value=round(s.total_deductions, 2))
        ws.cell(row=row, column=14, value=round(s.net_pay, 2))
        style_data_row(ws, row, len(headers))

        totals['gross'] += s.gross
        totals['pf']    += s.pf_employee
        totals['esic']  += s.esic_employee
        totals['ded']   += s.total_deductions
        totals['net']   += s.net_pay

    # Totals row
    trow = len(salaries) + 3
    ws.cell(row=trow, column=1,  value='TOTAL')
    ws.cell(row=trow, column=10, value=round(totals['gross'], 2))
    ws.cell(row=trow, column=11, value=round(totals['pf'],    2))
    ws.cell(row=trow, column=12, value=round(totals['esic'],  2))
    ws.cell(row=trow, column=13, value=round(totals['ded'],   2))
    ws.cell(row=trow, column=14, value=round(totals['net'],   2))
    style_total_row(ws, trow, len(headers))

    auto_width(ws)
    ws.freeze_panes = 'A3'
    return to_stream(wb)


# ── PF / ESIC Compliance Excel ─────────────────────────────────────────
def export_compliance_to_excel(records, month, year):
    wb = Workbook()
    ws = wb.active
    ws.title = f'Compliance {month:02d}-{year}'

    ws.merge_cells('A1:J1')
    ws['A1'] = f'PF & ESIC COMPLIANCE — {month:02d}/{year}'
    ws['A1'].font      = Font(bold=True, size=14, color='1A237E')
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 30

    headers = [
        '#', 'Worker Name', 'Company', 'Post',
        'PF (Emp 12%)', 'PF (Emplr 13%)', 'Total PF',
        'ESIC (Emp 0.75%)', 'ESIC (Emplr 3.25%)', 'Total ESIC'
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, len(headers))

    totals = {k: 0 for k in ['pf_e','pf_r','esic_e','esic_r']}
    for i, r in enumerate(records, 1):
        row = i + 2
        ws.cell(row=row, column=1,  value=i)
        ws.cell(row=row, column=2,  value=r.worker.full_name)
        ws.cell(row=row, column=3,  value=r.company.name)
        ws.cell(row=row, column=4,  value=r.worker.post)
        ws.cell(row=row, column=5,  value=round(r.pf_employee,   2))
        ws.cell(row=row, column=6,  value=round(r.pf_employer,   2))
        ws.cell(row=row, column=7,  value=round(r.pf_employee + r.pf_employer, 2))
        ws.cell(row=row, column=8,  value=round(r.esic_employee, 2))
        ws.cell(row=row, column=9,  value=round(r.esic_employer, 2))
        ws.cell(row=row, column=10, value=round(r.esic_employee + r.esic_employer, 2))
        style_data_row(ws, row, len(headers))

        totals['pf_e']   += r.pf_employee
        totals['pf_r']   += r.pf_employer
        totals['esic_e'] += r.esic_employee
        totals['esic_r'] += r.esic_employer

    trow = len(records) + 3
    ws.cell(row=trow, column=1,  value='TOTAL')
    ws.cell(row=trow, column=5,  value=round(totals['pf_e'],   2))
    ws.cell(row=trow, column=6,  value=round(totals['pf_r'],   2))
    ws.cell(row=trow, column=7,  value=round(totals['pf_e'] + totals['pf_r'], 2))
    ws.cell(row=trow, column=8,  value=round(totals['esic_e'], 2))
    ws.cell(row=trow, column=9,  value=round(totals['esic_r'], 2))
    ws.cell(row=trow, column=10, value=round(totals['esic_e'] + totals['esic_r'], 2))
    style_total_row(ws, trow, len(headers))

    auto_width(ws)
    ws.freeze_panes = 'A3'
    return to_stream(wb)


# ── Workers KYC Excel ──────────────────────────────────────────────────
def export_workers_to_excel(workers):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Workers KYC'

    ws.merge_cells('A1:K1')
    ws['A1'] = 'WORKERS & KYC REPORT'
    ws['A1'].font      = Font(bold=True, size=14, color='1A237E')
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 30

    headers = [
        '#', 'Full Name', 'Father Name', 'Post', 'Mobile',
        'Aadhaar', 'PAN', 'Bank Name', 'Account No',
        'IFSC', 'KYC Status'
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, len(headers))

    for i, w in enumerate(workers, 1):
        row = i + 2
        ws.cell(row=row, column=1,  value=i)
        ws.cell(row=row, column=2,  value=w.full_name)
        ws.cell(row=row, column=3,  value=w.father_name)
        ws.cell(row=row, column=4,  value=w.post)
        ws.cell(row=row, column=5,  value=w.mobile)
        ws.cell(row=row, column=6,  value=w.aadhaar_number)
        ws.cell(row=row, column=7,  value=w.pan_number)
        ws.cell(row=row, column=8,  value=w.bank_name)
        ws.cell(row=row, column=9,  value=w.account_number)
        ws.cell(row=row, column=10, value=w.ifsc_code)
        ws.cell(row=row, column=11,
                value='Complete' if w.kyc_complete() else 'Pending')
        style_data_row(ws, row, len(headers))

    auto_width(ws)
    ws.freeze_panes = 'A3'
    return to_stream(wb) 

def export_deployment_to_excel(requirements, month, year):
    from app.models.deployment import Deployment
    wb = Workbook()
    ws = wb.active
    ws.title = f'Deployment {month:02d}-{year}'

    ws.merge_cells('A1:H1')
    ws['A1'] = f'DEPLOYMENT REPORT — {month:02d}/{year}'
    ws['A1'].font      = Font(bold=True, size=14, color='1A237E')
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 30

    headers = [
        '#', 'Company', 'Post', 'Shift',
        'Required', 'Deployed', 'Difference', 'Status'
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, len(headers))

    for i, req in enumerate(requirements, 1):
        deployed = Deployment.query.filter_by(
            company_id=req.company_id,
            post=req.post,
            is_active=True
        ).count()
        diff   = deployed - req.required_count
        status = 'Matched' if diff == 0 else ('Short' if diff < 0 else 'Excess')

        row = i + 2
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=req.company.name)
        ws.cell(row=row, column=3, value=req.post)
        ws.cell(row=row, column=4, value=req.shift)
        ws.cell(row=row, column=5, value=req.required_count)
        ws.cell(row=row, column=6, value=deployed)
        ws.cell(row=row, column=7, value=diff)
        ws.cell(row=row, column=8, value=status)
        style_data_row(ws, row, len(headers))

    auto_width(ws)
    ws.freeze_panes = 'A3'
    return to_stream(wb)

def export_quarterly_to_excel(salaries, quarter, year):
    wb  = Workbook()
    ws  = wb.active
    ws.title = f'Q{quarter}-{year}'

    ws.merge_cells('A1:N1')
    ws['A1'] = f'QUARTERLY SALARY REPORT — Q{quarter} / {year}'
    ws['A1'].font      = Font(bold=True, size=14, color='1A237E')
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 30

    headers = [
        '#', 'Worker Name', 'Company', 'Post', 'Month',
        'Days', 'Basic', 'DA', 'HRA', 'Special',
        'Gross', 'PF', 'ESIC', 'Net Pay'
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, len(headers))

    totals = {k: 0 for k in ['gross','pf','esic','net']}
    for i, s in enumerate(salaries, 1):
        row = i + 2
        ws.cell(row=row, column=1,  value=i)
        ws.cell(row=row, column=2,  value=s.worker.full_name)
        ws.cell(row=row, column=3,  value=s.company.name)
        ws.cell(row=row, column=4,  value=s.worker.post)
        ws.cell(row=row, column=5,  value=f'{s.month:02d}/{s.year}')
        ws.cell(row=row, column=6,  value=s.days_present)
        ws.cell(row=row, column=7,  value=round(s.basic, 2))
        ws.cell(row=row, column=8,  value=round(s.da, 2))
        ws.cell(row=row, column=9,  value=round(s.hra, 2))
        ws.cell(row=row, column=10, value=round(s.special_allowance, 2))
        ws.cell(row=row, column=11, value=round(s.gross, 2))
        ws.cell(row=row, column=12, value=round(s.pf_employee, 2))
        ws.cell(row=row, column=13, value=round(s.esic_employee, 2))
        ws.cell(row=row, column=14, value=round(s.net_pay, 2))
        style_data_row(ws, row, len(headers))

        totals['gross'] += s.gross
        totals['pf']    += s.pf_employee
        totals['esic']  += s.esic_employee
        totals['net']   += s.net_pay

    trow = len(salaries) + 3
    ws.cell(row=trow, column=1,  value='TOTAL')
    ws.cell(row=trow, column=11, value=round(totals['gross'], 2))
    ws.cell(row=trow, column=12, value=round(totals['pf'],    2))
    ws.cell(row=trow, column=13, value=round(totals['esic'],  2))
    ws.cell(row=trow, column=14, value=round(totals['net'],   2))
    style_total_row(ws, trow, len(headers))

    auto_width(ws)
    ws.freeze_panes = 'A3'
    return to_stream(wb)

def export_payment_to_excel(salaries, month, year):
    wb  = Workbook()
    ws  = wb.active
    ws.title = f'Payment {month:02d}-{year}'

    ws.merge_cells('A1:K1')
    ws['A1'] = f'SALARY PAYMENT SHEET — {month:02d}/{year}'
    ws['A1'].font      = Font(bold=True, size=14,
                              color='FFFFFF')
    ws['A1'].fill      = PatternFill('solid', fgColor='1A237E')
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 30

    headers = [
        'Emp ID', 'Employee Name', 'Post', 'Company',
        'Bank Name', 'Account No', 'IFSC Code',
        'Gross (₹)', 'Deductions (₹)',
        'Net Pay (₹)', 'Status'
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, len(headers))
    ws.row_dimensions[2].height = 20

    total_net = 0
    for i, s in enumerate(salaries, 1):
        row = i + 2
        ws.cell(row=row, column=1,
                value=s.worker.employee_id or
                      f'EMP{s.worker.id:04d}')
        ws.cell(row=row, column=2,
                value=s.worker.full_name)
        ws.cell(row=row, column=3,
                value=s.worker.post)
        ws.cell(row=row, column=4,
                value=s.company.name)
        ws.cell(row=row, column=5,
                value=s.worker.bank_name or '—')
        ws.cell(row=row, column=6,
                value=s.worker.account_number or '—')
        ws.cell(row=row, column=7,
                value=s.worker.ifsc_code or '—')
        ws.cell(row=row, column=8,
                value=round(s.gross, 2))
        ws.cell(row=row, column=9,
                value=round(s.total_deductions, 2))
        ws.cell(row=row, column=10,
                value=round(s.net_pay, 2))
        ws.cell(row=row, column=11, value='Pending')
        style_data_row(ws, row, len(headers))
        total_net += s.net_pay

    # Total row
    trow = len(salaries) + 3
    ws.cell(row=trow, column=1,  value='TOTAL')
    ws.cell(row=trow, column=10,
            value=round(total_net, 2))
    ws.cell(row=trow, column=11,
            value=f'{len(salaries)} Employees')
    style_total_row(ws, trow, len(headers))

    auto_width(ws)
    ws.freeze_panes = 'A3'
    return to_stream(wb)

def export_slips_to_excel(salaries, month, year):
    from app.utils.salary_engine import MONTH_NAMES
    wb = Workbook()

    for s in salaries:
        # One sheet per employee
        ws = wb.create_sheet(
            title=f'{s.worker.employee_id or s.worker.id}'
        )

        # Header
        ws.merge_cells('A1:D1')
        ws['A1'] = 'SALARY SLIP'
        ws['A1'].font      = Font(bold=True, size=14,
                                   color='FFFFFF')
        ws['A1'].fill      = PatternFill('solid',
                                          fgColor='1A237E')
        ws['A1'].alignment = CENTER
        ws.row_dimensions[1].height = 25

        ws.merge_cells('A2:D2')
        ws['A2'] = f'{MONTH_NAMES[month]} {year}'
        ws['A2'].font      = Font(bold=True, size=12,
                                   color='FFFFFF')
        ws['A2'].fill      = PatternFill('solid',
                                          fgColor='283593')
        ws['A2'].alignment = CENTER

        # Worker info
        info = [
            ('Employee Name', s.worker.full_name),
            ('Employee ID',
             s.worker.employee_id or
             f'EMP{s.worker.id:04d}'),
            ('Designation', s.worker.post or '—'),
            ('Company', s.company.name),
            ('Days Present', str(s.days_present)),
            ('Bank', s.worker.bank_name or '—'),
            ('Account No', s.worker.account_number or '—'),
        ]
        for i, (k, v) in enumerate(info, 4):
            ws.cell(row=i, column=1, value=k).font = \
                Font(bold=True)
            ws.cell(row=i, column=2, value=v)

        # Earnings
        row = 12
        ws.cell(row=row, column=1,
                value='EARNINGS').font = \
            Font(bold=True, color='FFFFFF')
        ws.cell(row=row, column=1).fill = \
            PatternFill('solid', fgColor='2E7D32')
        ws.cell(row=row, column=2,
                value='AMOUNT').font = \
            Font(bold=True, color='FFFFFF')
        ws.cell(row=row, column=2).fill = \
            PatternFill('solid', fgColor='2E7D32')
        ws.cell(row=row, column=3,
                value='DEDUCTIONS').font = \
            Font(bold=True, color='FFFFFF')
        ws.cell(row=row, column=3).fill = \
            PatternFill('solid', fgColor='C62828')
        ws.cell(row=row, column=4,
                value='AMOUNT').font = \
            Font(bold=True, color='FFFFFF')
        ws.cell(row=row, column=4).fill = \
            PatternFill('solid', fgColor='C62828')

        earnings = [
            ('Basic',           s.basic),
            ('DA',              s.da),
            ('HRA',             s.hra),
            ('Special Allowance', s.special_allowance),
        ]
        deductions = [
            ('PF (Employee 12%)', s.pf_employee),
            ('ESIC (Employee 0.75%)', s.esic_employee),
            ('Advance',          s.advance),
            ('Other',            s.other_deductions),
        ]

        max_rows = max(len(earnings), len(deductions))
        for j in range(max_rows):
            r = row + j + 1
            if j < len(earnings):
                ws.cell(row=r, column=1,
                        value=earnings[j][0])
                ws.cell(row=r, column=2,
                        value=round(earnings[j][1], 2))
            if j < len(deductions):
                ws.cell(row=r, column=3,
                        value=deductions[j][0])
                ws.cell(row=r, column=4,
                        value=round(deductions[j][1], 2))

        # Gross and Net
        gross_row = row + max_rows + 1
        ws.cell(row=gross_row, column=1,
                value='GROSS SALARY')
        ws.cell(row=gross_row, column=1).font = \
            Font(bold=True)
        ws.cell(row=gross_row, column=2,
                value=round(s.gross, 2)).font = \
            Font(bold=True, color='1B5E20')

        ws.cell(row=gross_row, column=3,
                value='TOTAL DEDUCTIONS')
        ws.cell(row=gross_row, column=3).font = \
            Font(bold=True)
        ws.cell(row=gross_row, column=4,
                value=round(s.total_deductions, 2)).font = \
            Font(bold=True, color='B71C1C')

        net_row = gross_row + 1
        ws.merge_cells(
            f'A{net_row}:B{net_row}'
        )
        ws.cell(row=net_row, column=1,
                value='NET PAY').font = \
            Font(bold=True, size=13, color='FFFFFF')
        ws.cell(row=net_row, column=1).fill = \
            PatternFill('solid', fgColor='1A237E')
        ws.cell(row=net_row, column=1).alignment = CENTER

        ws.merge_cells(
            f'C{net_row}:D{net_row}'
        )
        ws.cell(row=net_row, column=3,
                value=f'Rs. {round(s.net_pay, 2)}').font = \
            Font(bold=True, size=13, color='FFFFFF')
        ws.cell(row=net_row, column=3).fill = \
            PatternFill('solid', fgColor='1A237E')
        ws.cell(row=net_row, column=3).alignment = CENTER

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15

    # Remove default empty sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    return to_stream(wb)
