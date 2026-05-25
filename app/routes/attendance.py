from flask import (Blueprint, render_template, request,
                   redirect, url_for, flash, send_file)
from app import db
from app.models.client import Company
from app.models.worker import Worker
from app.models.deployment import Deployment
from app.models.attendance import Attendance
from datetime import date
from app.routes.auth import (login_required,
                              permission_required)
import calendar
import io

attendance_bp = Blueprint('attendance', __name__)

MONTH_NAMES = {
    1:'January',  2:'February', 3:'March',
    4:'April',    5:'May',      6:'June',
    7:'July',     8:'August',   9:'September',
    10:'October', 11:'November', 12:'December'
}


@attendance_bp.route('/attendance')
@login_required
@permission_required('attendance', 'view')
def index():
    month      = int(request.args.get('month', date.today().month))
    year       = int(request.args.get('year',  date.today().year))
    company_id = request.args.get('company_id', '')
    companies  = Company.query.all()

    query = Deployment.query.filter_by(is_active=True)
    if company_id:
        query = query.filter_by(company_id=company_id)
    deployments = query.all()

    days_in_month = calendar.monthrange(year, month)[1]
    summary = []
    for dep in deployments:
        records = Attendance.query.filter_by(
            worker_id=dep.worker_id
        ).filter(
            db.func.strftime('%m', Attendance.date) == '%02d' % month,
            db.func.strftime('%Y', Attendance.date) == str(year)
        ).all()

        present = sum(1 for r in records if r.status == 'P')
        absent  = sum(1 for r in records if r.status == 'A')
        halfday = sum(1 for r in records if r.status == 'H')
        marked  = len(records)

        summary.append({
            'worker'       : dep.worker,
            'company'      : dep.company,
            'post'         : dep.post,
            'present'      : present,
            'absent'       : absent,
            'halfday'      : halfday,
            'marked'       : marked,
            'days_in_month': days_in_month,
            'pending'      : days_in_month - marked
        })

    return render_template('attendance/index.html',
                           summary=summary,
                           companies=companies,
                           month=month, year=year,
                           month_name=MONTH_NAMES[month],
                           company_id=company_id)


@attendance_bp.route('/attendance/mark', methods=['GET', 'POST'])
@login_required
@permission_required('attendance', 'add')
def mark():
    companies  = Company.query.all()
    month      = int(request.args.get('month', date.today().month))
    year       = int(request.args.get('year',  date.today().year))
    company_id = request.args.get('company_id', '')

    days_in_month = calendar.monthrange(year, month)[1]
    days          = list(range(1, days_in_month + 1))

    # Get day names and weekend flags
    day_info = []
    for d in days:
        dt       = date(year, month, d)
        weekday  = dt.weekday()  # 0=Mon, 5=Sat, 6=Sun
        day_info.append({
            'day'       : d,
            'name'      : dt.strftime('%a'),
            'is_weekend': weekday >= 5
        })

    query = Deployment.query.filter_by(is_active=True)
    if company_id:
        query = query.filter_by(company_id=int(company_id))
    deployments = query.all()

    # Build attendance grid
    attendance_grid = {}
    for dep in deployments:
        records = Attendance.query.filter_by(
            worker_id=dep.worker_id
        ).filter(
            db.func.strftime('%m', Attendance.date) == '%02d' % month,
            db.func.strftime('%Y', Attendance.date) == str(year)
        ).all()
        attendance_grid[dep.worker_id] = {
            r.date.day: r.status for r in records
        }

    if request.method == 'POST':
        month      = int(request.form['month'])
        year       = int(request.form['year'])
        company_id = request.form.get('company_id', '')
        worker_ids = request.form.getlist('worker_ids')

        for worker_id in worker_ids:
            worker_id = int(worker_id)
            for d in range(1, calendar.monthrange(year, month)[1] + 1):
                field_name = f'att_{worker_id}_{d}'
                status     = request.form.get(field_name, 'A')
                att_date   = date(year, month, d)

                existing = Attendance.query.filter_by(
                    worker_id=worker_id,
                    date=att_date
                ).first()
                if existing:
                    existing.status = status
                else:
                    db.session.add(Attendance(
                        worker_id=worker_id,
                        date=att_date,
                        status=status
                    ))

        db.session.commit()
        flash('Attendance saved successfully!', 'success')
        return redirect(url_for('attendance.index',
                                month=month, year=year,
                                company_id=company_id))

    return render_template('attendance/mark.html',
                           deployments=deployments,
                           companies=companies,
                           day_info=day_info,
                           days=days,
                           month=month, year=year,
                           month_name=MONTH_NAMES[month],
                           company_id=company_id,
                           attendance_grid=attendance_grid,
                           days_in_month=days_in_month)


@attendance_bp.route('/attendance/download-template')
def download_template():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    month      = int(request.args.get('month', date.today().month))
    year       = int(request.args.get('year',  date.today().year))
    company_id = request.args.get('company_id', '')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance'

    days_in_month = calendar.monthrange(year, month)[1]

    # Header
    ws.merge_cells(f'A1:{get_column_letter(days_in_month + 3)}1')
    ws['A1'] = f'ATTENDANCE SHEET — {MONTH_NAMES[month]} {year}'
    ws['A1'].font      = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill      = PatternFill('solid', fgColor='1A237E')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    ws['A2'] = 'Employee ID'
    ws['B2'] = 'Employee Name'
    ws['C2'] = 'Post'

    weekend_fill = PatternFill('solid', fgColor='FFE0B2')
    header_fill  = PatternFill('solid', fgColor='E8EAF6')

    for d in range(1, days_in_month + 1):
        col  = d + 3
        cell = ws.cell(row=2, column=col)
        dt   = date(year, month, d)
        cell.value     = f'{d}\n{dt.strftime("%a")}'
        cell.alignment = Alignment(
            horizontal='center', wrap_text=True)
        cell.font = Font(bold=True, size=9)
        if dt.weekday() >= 5:
            cell.fill = weekend_fill
        else:
            cell.fill = header_fill
        ws.column_dimensions[
            get_column_letter(col)].width = 5

    ws.row_dimensions[2].height = 30

    # Get deployed workers
    query = Deployment.query.filter_by(is_active=True)
    if company_id:
        query = query.filter_by(company_id=int(company_id))
    deployments = query.all()

    # Fill worker rows
    for i, dep in enumerate(deployments):
        row = i + 3
        ws.cell(row=row, column=1,
                value=dep.worker.employee_id or f'EMP{dep.worker.id:04d}')
        ws.cell(row=row, column=2,
                value=dep.worker.full_name)
        ws.cell(row=row, column=3,
                value=dep.post)

        for d in range(1, days_in_month + 1):
            col  = d + 3
            cell = ws.cell(row=row, column=col)
            dt   = date(year, month, d)
            if dt.weekday() >= 5:
                cell.value = 'WO'
                cell.fill  = weekend_fill
            else:
                cell.value = 'P'
            cell.alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 15

    # Instructions sheet
    ws2 = wb.create_sheet('Instructions')
    instructions = [
        ['Attendance Code', 'Meaning'],
        ['P',  'Present'],
        ['A',  'Absent'],
        ['H',  'Half Day'],
        ['WO', 'Week Off (Sunday/Saturday)'],
        ['PH', 'Public Holiday'],
        ['', ''],
        ['IMPORTANT:', 'Do not change Employee ID or Name'],
        ['IMPORTANT:', 'Use only the codes listed above'],
        ['IMPORTANT:', 'Do not add or remove columns'],
    ]
    for row in instructions:
        ws2.append(row)
    ws2['A1'].font = Font(bold=True)
    ws2['B1'].font = Font(bold=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Attendance_Template_{MONTH_NAMES[month]}_{year}.xlsx'
    )


@attendance_bp.route('/attendance/import-excel', methods=['POST'])
def import_excel():
    from openpyxl import load_workbook
    from app.models.worker import Worker
    import re

    # Get form data
    month = int(request.form['month'])
    year = int(request.form['year'])
    company_id = request.form.get('company_id', '')
    file = request.files.get('excel_file')

    if not file:
        flash('Please select an Excel file!', 'danger')
        return redirect(url_for('attendance.index'))

    # Check if file is valid
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('Invalid file format! Please upload .xlsx or .xls file only.', 'danger')
        return redirect(url_for('attendance.index'))

    try:
        wb = load_workbook(file)
        ws = wb.active
        
        # Try to read month/year from the template header (cell A1)
        header_text = str(ws['A1'].value or '')
        
        # Extract month and year from header like "ATTENDANCE SHEET — March 2025"
        month_match = re.search(r'([A-Za-z]+)\s+(\d{4})', header_text)
        if month_match:
            month_name = month_match.group(1)
            year_from_file = int(month_match.group(2))
            
            # Convert month name to number
            month_names = {
                'January': 1, 'February': 2, 'March': 3, 'April': 4,
                'May': 5, 'June': 6, 'July': 7, 'August': 8,
                'September': 9, 'October': 10, 'November': 11, 'December': 12
            }
            
            if month_name in month_names:
                month = month_names[month_name]
                year = year_from_file
                flash(f'Using month/year from file: {month_name} {year}', 'info')
        
        days_in_month = calendar.monthrange(year, month)[1]
        saved = 0
        updated = 0
        errors = []

        # Process rows (starting from row 3 as per template)
        for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=False), start=3):
            emp_id_cell = row[0]
            emp_id = str(emp_id_cell.value or '').strip() if emp_id_cell else ''
            
            if not emp_id:
                continue

            # Find worker by employee_id
            worker = Worker.query.filter_by(employee_id=emp_id).first()
            if not worker:
                errors.append(f'Row {row_num}: Employee ID "{emp_id}" not found')
                continue

            # Process each day column (starting from column D = index 3)
            for d in range(1, days_in_month + 1):
                col_index = d + 2  # A=0, B=1, C=2, D=3 -> day 1
                if col_index >= len(row):
                    break
                    
                status_cell = row[col_index]
                status = str(status_cell.value or 'A').strip().upper()
                
                # Map WO and PH to A for calculation
                if status in ['WO', 'PH']:
                    status = 'A'
                if status not in ['P', 'A', 'H']:
                    status = 'A'

                att_date = date(year, month, d)
                existing = Attendance.query.filter_by(
                    worker_id=worker.id,
                    date=att_date
                ).first()
                
                if existing:
                    existing.status = status
                    updated += 1
                else:
                    db.session.add(Attendance(
                        worker_id=worker.id,
                        date=att_date,
                        status=status
                    ))
                    saved += 1

        db.session.commit()

        # Flash summary message
        total = saved + updated
        if errors:
            flash(f'Import completed with {len(errors)} errors: {", ".join(errors[:3])}', 'warning')
        else:
            flash(f'Successfully imported {total} attendance records for {month}/{year}! (New: {saved}, Updated: {updated})', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error importing file: {str(e)}', 'danger')

    return redirect(url_for('attendance.index', month=month, year=year, company_id=company_id))


@attendance_bp.route('/attendance/export-excel')
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    month      = int(request.args.get('month', date.today().month))
    year       = int(request.args.get('year',  date.today().year))
    company_id = request.args.get('company_id', '')

    wb = Workbook()
    ws = wb.active
    ws.title = f'{MONTH_NAMES[month][:3]}-{year}'

    days_in_month = calendar.monthrange(year, month)[1]

    # Header
    total_cols = days_in_month + 7
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    ws['A1'] = f'ATTENDANCE REGISTER — {MONTH_NAMES[month].upper()} {year}'
    ws['A1'].font      = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill      = PatternFill('solid', fgColor='1A237E')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 25

    # Column headers
    headers = ['Emp ID', 'Name', 'Post']
    ws['A2'] = headers[0]
    ws['B2'] = headers[1]
    ws['C2'] = headers[2]

    weekend_fill = PatternFill('solid', fgColor='FFE0B2')
    header_fill  = PatternFill('solid', fgColor='E8EAF6')
    present_fill = PatternFill('solid', fgColor='C8E6C9')
    absent_fill  = PatternFill('solid', fgColor='FFCDD2')
    half_fill    = PatternFill('solid', fgColor='FFF9C4')

    for d in range(1, days_in_month + 1):
        col  = d + 3
        cell = ws.cell(row=2, column=col)
        dt   = date(year, month, d)
        cell.value     = f'{d}\n{dt.strftime("%a")}'
        cell.alignment = Alignment(
            horizontal='center', wrap_text=True)
        cell.font = Font(bold=True, size=9)
        cell.fill = weekend_fill if dt.weekday() >= 5 \
                    else header_fill
        ws.column_dimensions[
            get_column_letter(col)].width = 5

    # Summary headers
    sum_col = days_in_month + 4
    for i, h in enumerate(['P', 'A', 'H', 'Total']):
        cell       = ws.cell(row=2, column=sum_col + i)
        cell.value = h
        cell.font  = Font(bold=True, size=10)
        cell.fill  = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[
            get_column_letter(sum_col + i)].width = 6

    ws.row_dimensions[2].height = 30

    # Get deployments
    query = Deployment.query.filter_by(is_active=True)
    if company_id:
        query = query.filter_by(company_id=int(company_id))
    deployments = query.all()

    # Fill data
    for i, dep in enumerate(deployments):
        row    = i + 3
        worker = dep.worker

        ws.cell(row=row, column=1,
                value=worker.employee_id or f'EMP{worker.id:04d}')
        ws.cell(row=row, column=2, value=worker.full_name)
        ws.cell(row=row, column=3, value=dep.post)

        records = Attendance.query.filter_by(
            worker_id=worker.id
        ).filter(
            db.func.strftime('%m', Attendance.date) == '%02d' % month,
            db.func.strftime('%Y', Attendance.date) == str(year)
        ).all()
        att_map = {r.date.day: r.status for r in records}

        p_count = a_count = h_count = 0
        for d in range(1, days_in_month + 1):
            col    = d + 3
            cell   = ws.cell(row=row, column=col)
            dt     = date(year, month, d)
            status = att_map.get(d, '')

            if dt.weekday() >= 5:
                cell.value = 'WO'
                cell.fill  = weekend_fill
            elif status == 'P':
                cell.value = 'P'
                cell.fill  = present_fill
                p_count   += 1
            elif status == 'H':
                cell.value = 'H'
                cell.fill  = half_fill
                h_count   += 1
            elif status == 'A':
                cell.value = 'A'
                cell.fill  = absent_fill
                a_count   += 1
            else:
                cell.value = '—'

            cell.alignment = Alignment(horizontal='center')
            cell.font      = Font(size=9)

        # Summary
        ws.cell(row=row, column=sum_col,     value=p_count)
        ws.cell(row=row, column=sum_col + 1, value=a_count)
        ws.cell(row=row, column=sum_col + 2, value=h_count)
        ws.cell(row=row, column=sum_col + 3,
                value=p_count + a_count + h_count)

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 15
    ws.freeze_panes                 = 'D3'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Attendance_{MONTH_NAMES[month]}_{year}.xlsx'
    )
