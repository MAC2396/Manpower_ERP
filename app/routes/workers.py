from flask import Blueprint, render_template, request, redirect, url_for, flash
import os
from werkzeug.utils import secure_filename
from datetime import datetime
from app import db
from app.routes.auth import (login_required,
                              permission_required)

workers_bp = Blueprint('workers', __name__)

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'pdf'}

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_file(file, folder, prefix):
    from flask import current_app
    if file and allowed_file(file.filename):
        filename  = secure_filename(file.filename)
        ext       = filename.rsplit('.', 1)[1].lower()
        new_name  = f'{prefix}_{datetime.now().strftime("%Y%m%d%H%M%S")}.{ext}'
        save_path = os.path.join(current_app.config['UPLOAD_FOLDER'], folder)
        os.makedirs(save_path, exist_ok=True)
        file.save(os.path.join(save_path, new_name))
        return os.path.join(folder, new_name)
    return None


@workers_bp.route('/workers')
@login_required
@permission_required('workers', 'view')
def index():
    from app.models.worker import Worker
    from app.models.client import Company
    from app.models.deployment import Deployment

    companies  = Company.query.all()

    # Get filter parameters
    search     = request.args.get('search', '').strip()
    company_id = request.args.get('company_id', '')
    kyc_filter = request.args.get('kyc_filter', '')

    # Base query
    query = Worker.query

    # Search by name or employee ID
    if search:
        query = query.filter(
            db.or_(
                Worker.full_name.ilike(f'%{search}%'),
                Worker.employee_id.ilike(f'%{search}%'),
                Worker.mobile.ilike(f'%{search}%')
            )
        )

    # Filter by KYC status
    if kyc_filter == 'pending':
        workers = [w for w in query.all()
                   if not w.kyc_complete()]
    elif kyc_filter == 'complete':
        workers = [w for w in query.all()
                   if w.kyc_complete()]
    else:
        workers = query.all()

    # Filter by company
    if company_id:
        deployed_ids = [
            d.worker_id for d in
            Deployment.query.filter_by(
                company_id=int(company_id),
                is_active=True
            ).all()
        ]
        workers = [w for w in workers
                   if w.id in deployed_ids]

    # Get company for each worker
    worker_company = {}
    for w in workers:
        dep = Deployment.query.filter_by(
            worker_id=w.id,
            is_active=True
        ).first()
        worker_company[w.id] = dep.company.name \
            if dep else '—'

    # KYC stats
    total_workers   = Worker.query.count()
    kyc_complete    = sum(
         1 for w in all_workers if w.kyc_complete()
    )
    kyc_pending     = total_workers - kyc_complete

    return render_template('workers/index.html',
                           workers=workers,
                           companies=companies,
                           worker_company=worker_company,
                           search=search,
                           company_id=company_id,
                           kyc_filter=kyc_filter,
                           total_workers=total_workers,
                           kyc_complete=kyc_complete,
                           kyc_pending=kyc_pending)


@workers_bp.route('/workers/add', methods=['GET', 'POST'])
@login_required
@permission_required('workers', 'add')
def add():
    from app import db
    from app.models.worker import Worker
    from app.models.client import Company
    companies = Company.query.all()

    if request.method == 'POST':
        worker = Worker(
            full_name      = request.form['full_name'],
            father_name    = request.form['father_name'],
            mobile         = request.form['mobile'],
            address        = request.form['address'],
            post           = request.form['post'],
            gender         = request.form.get('gender'),
            aadhaar_number = request.form['aadhaar_number'],
            pan_number     = request.form['pan_number'],
            bank_name      = request.form['bank_name'],
            account_number = request.form['account_number'],
            ifsc_code      = request.form['ifsc_code']
        )

        # Handle file uploads
        worker.photo         = save_file(
            request.files.get('photo'), 'photos', 'photo')
        worker.aadhaar_doc   = save_file(
            request.files.get('aadhaar_doc'), 'aadhaar', 'aadhaar')
        worker.pan_doc       = save_file(
            request.files.get('pan_doc'), 'pan', 'pan')
        worker.bank_passbook = save_file(
            request.files.get('bank_passbook'), 'bank', 'bank')

        db.session.add(worker)
        db.session.flush()

        # Auto generate employee ID
        worker.employee_id = f'EMP{worker.id:04d}'
        db.session.commit()

        flash(f'Employee added! ID: {worker.employee_id}', 'success')
        return redirect(url_for('workers.index'))

    return render_template('workers/add.html',
                           companies=companies)


@workers_bp.route('/workers/<int:id>')
def view(id):
    from app.models.worker import Worker
    worker = Worker.query.get_or_404(id)
    return render_template('workers/view.html', worker=worker)


@workers_bp.route('/workers/<int:id>/upload-kyc', methods=['GET', 'POST'])
def upload_kyc(id):
    from app import db
    from app.models.worker import Worker
    worker = Worker.query.get_or_404(id)

    if request.method == 'POST':
        # Update KYC numbers
        worker.aadhaar_number = request.form.get(
            'aadhaar_number', worker.aadhaar_number)
        worker.pan_number     = request.form.get(
            'pan_number', worker.pan_number)
        worker.bank_name      = request.form.get(
            'bank_name', worker.bank_name)
        worker.account_number = request.form.get(
            'account_number', worker.account_number)
        worker.ifsc_code      = request.form.get(
            'ifsc_code', worker.ifsc_code)

        # Update files if uploaded
        photo = save_file(
            request.files.get('photo'), 'photos', 'photo')
        if photo:
            worker.photo = photo

        aadhaar_doc = save_file(
            request.files.get('aadhaar_doc'), 'aadhaar', 'aadhaar')
        if aadhaar_doc:
            worker.aadhaar_doc = aadhaar_doc

        pan_doc = save_file(
            request.files.get('pan_doc'), 'pan', 'pan')
        if pan_doc:
            worker.pan_doc = pan_doc

        bank_passbook = save_file(
            request.files.get('bank_passbook'), 'bank', 'bank')
        if bank_passbook:
            worker.bank_passbook = bank_passbook

        db.session.commit()
        flash('KYC documents updated successfully!', 'success')
        return redirect(url_for('workers.view', id=worker.id))

    return render_template('workers/upload_kyc.html', worker=worker)


@workers_bp.route('/workers/<int:id>/family', methods=['GET', 'POST'])
def family(id):
    from app import db
    from app.models.worker import Worker, FamilyMember
    from datetime import datetime
    worker = Worker.query.get_or_404(id)

    if request.method == 'POST':
        dob_str = request.form.get('date_of_birth')
        dob     = datetime.strptime(dob_str, '%Y-%m-%d').date() \
                  if dob_str else None
        member  = FamilyMember(
            worker_id     = worker.id,
            name          = request.form['name'],
            relation      = request.form['relation'],
            date_of_birth = dob,
            mobile        = request.form.get('mobile')
        )
        db.session.add(member)
        db.session.commit()
        flash('Family member added!', 'success')
        return redirect(url_for('workers.family', id=id))

    return render_template('workers/family.html', worker=worker)


@workers_bp.route('/workers/family/delete/<int:fid>')
def delete_family(fid):
    from app import db
    from app.models.worker import FamilyMember
    member = FamilyMember.query.get_or_404(fid)
    worker_id = member.worker_id
    db.session.delete(member)
    db.session.commit()
    flash('Family member removed.', 'info')
    return redirect(url_for('workers.family', id=worker_id))


@workers_bp.route('/workers/<int:id>/edit', methods=['GET', 'POST'])
def edit(id):
    from app import db
    from app.models.worker import Worker
    worker = Worker.query.get_or_404(id)

    if request.method == 'POST':
        worker.full_name      = request.form['full_name']
        worker.father_name    = request.form['father_name']
        worker.mobile         = request.form['mobile']
        worker.address        = request.form['address']
        worker.post           = request.form['post']
        worker.gender         = request.form['gender']
        worker.aadhaar_number = request.form['aadhaar_number']
        worker.pan_number     = request.form['pan_number']
        worker.bank_name      = request.form['bank_name']
        worker.account_number = request.form['account_number']
        worker.ifsc_code      = request.form['ifsc_code']

        # Handle date of birth
        from datetime import datetime
        dob = request.form.get('date_of_birth')
        if dob:
            worker.date_of_birth = datetime.strptime(dob, '%Y-%m-%d').date()

        db.session.commit()
        flash('Worker details updated successfully!', 'success')
        return redirect(url_for('workers.view', id=worker.id))

    return render_template('workers/edit.html', worker=worker)

@workers_bp.route('/workers/download-template')
def download_template():
    from openpyxl import Workbook
    from openpyxl.styles import (PatternFill, Font,
                                  Alignment)
    from openpyxl.utils import get_column_letter
    from flask import send_file
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = 'Employee Import'

    # Header style
    header_fill = PatternFill('solid', fgColor='1A237E')
    header_font = Font(bold=True, color='FFFFFF',
                       size=11)
    center      = Alignment(horizontal='center',
                             vertical='center')

    # Column definitions
    columns = [
        ('Full Name *',       25),
        ('Father Name',       20),
        ('Mobile',            15),
        ('Gender',            10),
        ('Date of Birth',     15),
        ('Post/Designation',  18),
        ('Address',           30),
        ('Aadhaar Number',    18),
        ('PAN Number',        15),
        ('Bank Name',         20),
        ('Account Number',    20),
        ('IFSC Code',         15),
    ]

    # Title row
    ws.merge_cells(
        f'A1:{get_column_letter(len(columns))}1'
    )
    ws['A1'] = 'EMPLOYEE BULK IMPORT TEMPLATE'
    ws['A1'].font      = Font(bold=True, size=13,
                               color='FFFFFF')
    ws['A1'].fill      = PatternFill('solid',
                                      fgColor='0D1757')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 28

    # Headers
    for col, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col, value=name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        ws.column_dimensions[
            get_column_letter(col)].width = width
    ws.row_dimensions[2].height = 22

    # Sample data rows
    samples = [
        ['Ramesh Kumar', 'Suresh Kumar',
         '9876543210', 'Male', '01/01/1990',
         'Helper', 'Plot 12, Pune',
         '1234 5678 9012', 'ABCPK1234D',
         'SBI', '12345678901', 'SBIN0001234'],
        ['Suresh Sharma', 'Mahesh Sharma',
         '9823456789', 'Male', '15/06/1988',
         'Supervisor', 'Flat 5, Mumbai',
         '9876 5432 1098', 'XYZPS5678E',
         'HDFC', '98765432101', 'HDFC0001234'],
    ]

    note_fill   = PatternFill('solid', fgColor='FFF9C4')
    sample_fill = PatternFill('solid', fgColor='E8F5E9')

    for r, row in enumerate(samples, 3):
        for c, val in enumerate(row, 1):
            cell       = ws.cell(row=r, column=c,
                                  value=val)
            cell.fill  = sample_fill
            cell.font  = Font(italic=True,
                               color='1B5E20')

    # Instructions sheet
    ws2 = wb.create_sheet('Instructions')
    instructions = [
        ['INSTRUCTIONS FOR BULK IMPORT', ''],
        ['', ''],
        ['Column', 'Rules'],
        ['Full Name *',    'Required. Cannot be empty.'],
        ['Father Name',    'Optional'],
        ['Mobile',         '10 digit mobile number'],
        ['Gender',         'Male / Female / Other'],
        ['Date of Birth',  'Format: DD/MM/YYYY'],
        ['Post/Designation','e.g. Helper, Supervisor'],
        ['Address',        'Full address'],
        ['Aadhaar Number', '12 digit Aadhaar number'],
        ['PAN Number',     '10 character PAN'],
        ['Bank Name',      'Full bank name'],
        ['Account Number', 'Bank account number'],
        ['IFSC Code',      '11 character IFSC code'],
        ['', ''],
        ['IMPORTANT NOTES', ''],
        ['1', 'Do not change column headers'],
        ['2', 'Delete sample rows before uploading'],
        ['3', 'Date format must be DD/MM/YYYY'],
        ['4', 'Duplicate Aadhaar or PAN will be skipped'],
        ['5', 'Maximum 500 rows per upload'],
    ]

    title_font  = Font(bold=True, size=13,
                       color='1A237E')
    header_font2 = Font(bold=True, color='FFFFFF')

    for r, row in enumerate(instructions, 1):
        for c, val in enumerate(row, 1):
            cell       = ws2.cell(row=r, column=c,
                                   value=val)
            cell.font  = Font(size=11)

    ws2['A1'].font = title_font
    ws2['A3'].font = header_font2
    ws2['A3'].fill = PatternFill('solid',
                                  fgColor='1A237E')
    ws2['B3'].font = header_font2
    ws2['B3'].fill = PatternFill('solid',
                                  fgColor='1A237E')
    ws2['A17'].font = Font(bold=True, color='B71C1C')
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 45

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import send_file
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Employee_Import_Template.xlsx'
    )


@workers_bp.route('/workers/bulk-import',
                  methods=['POST'])
def bulk_import():
    from app import db
    from app.models.worker import Worker
    from openpyxl import load_workbook
    from datetime import datetime
    from flask import send_file
    import io

    file = request.files.get('excel_file')
    if not file:
        flash('Please select an Excel file!', 'danger')
        return redirect(url_for('workers.index'))

    # Validate file extension
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('Invalid file format! Please upload '
              '.xlsx or .xls file only.', 'danger')
        return redirect(url_for('workers.index'))

    try:
        wb = load_workbook(file)
    except Exception:
        flash('Could not read the file. Please use '
              'the provided template.', 'danger')
        return redirect(url_for('workers.index'))

    # Validate sheet name
    if 'Employee Import' not in wb.sheetnames:
        flash(
            'Invalid template format! '
            'Sheet "Employee Import" not found. '
            'Please download and use the official '
            'template provided.',
            'danger'
        )
        return redirect(url_for('workers.index'))

    ws = wb['Employee Import']

    # Validate headers
    expected_headers = [
        'Full Name *', 'Father Name', 'Mobile',
        'Gender', 'Date of Birth', 'Post/Designation',
        'Address', 'Aadhaar Number', 'PAN Number',
        'Bank Name', 'Account Number', 'IFSC Code'
    ]
    actual_headers = [
        str(ws.cell(row=2, column=c).value or '').strip()
        for c in range(1, 13)
    ]

    if actual_headers != expected_headers:
        flash(
            'Column headers do not match the template! '
            'Please download and use the official '
            'template without changing column names.',
            'danger'
        )
        return redirect(url_for('workers.index'))

    # Process rows
    added   = 0
    skipped = 0
    errors  = []

    for row_num in range(3, ws.max_row + 1):
        full_name = str(
            ws.cell(row=row_num, column=1).value or ''
        ).strip()

        # Skip empty rows and sample rows
        if not full_name or full_name == 'Ramesh Kumar':
            continue

        # Skip only if name matches sample data exactly
        if full_name in ['Ramesh Kumar', 'Suresh Sharma']:
            continue

        try:
            # Parse date of birth
            dob_str = str(
                ws.cell(row=row_num, column=5).value
                or ''
            ).strip()
            dob = None
            if dob_str and dob_str != 'None':
                try:
                    dob = datetime.strptime(
                        dob_str, '%d/%m/%Y'
                    ).date()
                except ValueError:
                    try:
                        dob = datetime.strptime(
                            dob_str, '%Y-%m-%d'
                        ).date()
                    except ValueError:
                        pass

            aadhaar = str(
                ws.cell(row=row_num, column=8).value
                or ''
            ).strip().replace(' ', '')
            pan     = str(
                ws.cell(row=row_num, column=9).value
                or ''
            ).strip().upper()

            # Check duplicates
            if aadhaar:
                existing = Worker.query.filter_by(
                    aadhaar_number=aadhaar
                ).first()
                if existing:
                    skipped += 1
                    errors.append(
                        f'Row {row_num}: '
                        f'Aadhaar {aadhaar} already '
                        f'exists for '
                        f'{existing.full_name}'
                    )
                    continue

            # Create worker
            worker = Worker(
                full_name      = full_name,
                father_name    = str(
                    ws.cell(row=row_num,
                             column=2).value or ''
                ).strip() or None,
                mobile         = str(
                    ws.cell(row=row_num,
                             column=3).value or ''
                ).strip() or None,
                gender         = str(
                    ws.cell(row=row_num,
                             column=4).value or ''
                ).strip() or None,
                date_of_birth  = dob,
                post           = str(
                    ws.cell(row=row_num,
                             column=6).value or ''
                ).strip() or None,
                address        = str(
                    ws.cell(row=row_num,
                             column=7).value or ''
                ).strip() or None,
                aadhaar_number = aadhaar or None,
                pan_number     = pan or None,
                bank_name      = str(
                    ws.cell(row=row_num,
                             column=10).value or ''
                ).strip() or None,
                account_number = str(
                    ws.cell(row=row_num,
                             column=11).value or ''
                ).strip() or None,
                ifsc_code      = str(
                    ws.cell(row=row_num,
                             column=12).value or ''
                ).strip() or None,
            )

            db.session.add(worker)
            db.session.flush()

            # Generate Employee ID
            worker.employee_id = f'EMP{worker.id:04d}'
            added += 1

        except Exception as e:
            errors.append(
                f'Row {row_num}: Error — {str(e)}'
            )
            skipped += 1
            continue

    db.session.commit()

    msg = f'Import complete! {added} employees added'
    if skipped > 0:
        msg += f', {skipped} skipped'
    if errors:
        msg += f'. Errors: {"; ".join(errors[:3])}'
        flash(msg, 'warning')
    else:
        flash(msg, 'success')

    return redirect(url_for('workers.index'))
